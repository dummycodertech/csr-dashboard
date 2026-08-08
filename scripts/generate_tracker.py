"""
Converts Demo_Onboarding_Tracker.xlsx → src/data/trackerData.js
Run: python scripts/generate_tracker.py
"""
import openpyxl, json, os

wb = openpyxl.load_workbook(
    r'C:\Users\kyaba\OneDrive\Desktop\Demo_Onboarding_Tracker.xlsx',
    data_only=True
)

def parse_tracker(ws):
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    cols = []
    cur_section = 'EMPLOYEE DETAILS'
    cur_main = None

    for i, (r2, r3, r4) in enumerate(zip(row2, row3, row4)):
        if r2 is not None and str(r2).strip():
            cur_section = str(r2).strip()
        if r3 is not None and str(r3).strip():
            cur_main = str(r3).strip()

        if r4 is not None and str(r4).strip():
            col_name = str(r4).strip()
        elif cur_main:
            col_name = cur_main
        else:
            col_name = f'Col_{i}'

        cols.append({'section': cur_section, 'name': col_name, 'index': i})

    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        try:
            sno = int(row[0])
        except (ValueError, TypeError):
            continue

        rec = {}
        for col in cols:
            i = col['index']
            val = row[i] if i < len(row) else None
            rec[col['name']] = str(val).strip() if val is not None else ''
        records.append(rec)

    # Deduplicate column names (keep section context)
    seen = {}
    unique_cols = []
    for col in cols:
        key = col['name']
        if key in seen:
            col = {**col, 'name': f"{col['name']} ({col['section'].title()})"}
        else:
            seen[key] = True
        unique_cols.append(col)

    return {'columns': unique_cols, 'records': records}

onrole  = parse_tracker(wb['ONROLE Onboarding Tracker'])
offrole = parse_tracker(wb['OFFROLE Onboarding Tracker'])

print(f"ONROLE:  {len(onrole['records'])} records, {len(onrole['columns'])} columns")
print(f"OFFROLE: {len(offrole['records'])} records, {len(offrole['columns'])} columns")

out_path = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'trackerData.js')
payload = json.dumps({'onrole': onrole, 'offrole': offrole}, indent=2, ensure_ascii=False)

js = f"""// Onboarding Tracker Data — converted from Demo_Onboarding_Tracker.xlsx
// Generated 2026-08-08 — Source: Demo_Onboarding_Tracker.xlsx (portfolio/demo data)
export const trackerData = {payload};
"""

with open(out_path, 'w', encoding='utf-8') as f:
    f.write(js)

print(f"Written to {out_path} ({os.path.getsize(out_path)//1024} KB)")
